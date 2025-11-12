"""
Enhanced Excel output generation with improved formatting and data handling.
"""
import pandas as pd
import logging
import io
import math
import openpyxl
from typing import List, Dict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from coordinate_extraction import validate_coords_list, polyline_length_from_coords, safe_development_length
from openpyxl.utils import get_column_letter

def guarantee_and_format_results(results):
    # results: list of dicts
    out = []
    for r in results:
        rr = dict(r)
        # burst numeric -> formatted
        burst = rr.get('burst_pressure_mpa')
        try:
            burst = float(burst) if burst is not None else None
        except Exception:
            burst = None
        rr['burst_pressure_mpa'] = burst
        rr['burst_pressure_formatted'] = f"{burst:.2f} MPa" if burst is not None else None

        # id tolerance formatted
        idn = rr.get('id_nominal_mm') or rr.get('id1') or rr.get('ID')
        idt = rr.get('id_tolerance_mm')
        try:
            if idn is not None:
                idn = float(idn)
        except Exception:
            idn = None
        rr['id_nominal_mm'] = idn
        rr['id_tolerance_mm'] = float(idt) if idt is not None else None
        if idn is not None:
            if rr['id_tolerance_mm'] is not None:
                rr['id_formatted'] = f"{idn:.2f} ± {rr['id_tolerance_mm']:.2f} mm"
            else:
                rr['id_formatted'] = f"{idn:.2f} mm"
        else:
            rr['id_formatted'] = None

        # thickness fallback computed if missing
        odn = rr.get('od_nominal_mm')
        try:
            if odn is not None:
                odn = float(odn)
        except Exception:
            odn = None
        rr['od_nominal_mm'] = odn

        provenance = rr.get('thickness_source')
        # Only compute thickness if not set by authoritative table
        if rr.get('thickness_mm') is None and odn is not None and idn is not None and provenance != 'TABLE-4/8 Grade-1 authoritative':
            try:
                rr['thickness_mm'] = round((odn - idn) / 2.0, 3)
                # Only set tolerance if not already set by table
                if rr.get('thickness_tolerance_mm') is None:
                    rr['thickness_tolerance_mm'] = 0.25
                logging.debug(f"Computed thickness from OD/ID for part {rr.get('part_number','?')}: {rr['thickness_mm']} mm (fallback)")
            except Exception:
                rr['thickness_mm'] = None
        elif provenance == 'TABLE-4/8 Grade-1 authoritative':
            logging.debug(f"Thickness for part {rr.get('part_number','?')} is authoritative from table: {rr['thickness_mm']} mm ± {rr.get('thickness_tolerance_mm')} mm")

        # final display fields used in Excel
        rr['id_display'] = rr['id_formatted'] or "N/A"
        rr['od_display'] = (f"{odn:.2f} ± {rr['od_tolerance_mm']:.2f} mm" if odn is not None and rr.get('od_tolerance_mm') is not None else ("N/A" if odn is None else f"{odn:.2f} mm"))
        rr['thickness_display'] = (f"{rr['thickness_mm']:.3f} ± {rr.get('thickness_tolerance_mm'):.2f} mm" if rr.get('thickness_mm') is not None else "N/A")
        rr['burst_display'] = rr['burst_pressure_formatted'] or "N/A"

        out.append(rr)
    return out

def format_tolerance(val, tol):
    """Return formatted tolerance string or None if both are missing."""
    try:
        if val is None and tol is None:
            return None
        if val is None:
            return f"± {tol:.2f} mm" if tol is not None else None
        if tol is None:
            return f"{float(val):.2f} mm"
        return f"{float(val):.2f} ± {float(tol):.2f} mm"
    except Exception:
        return None

def ensure_result_fields(result: Dict) -> Dict:
    """
    Ensure result dict has expected fields populated or reasonable fallbacks.
    Fields set/formatted:
      - id_nominal_mm, id_tolerance_mm, id_formatted
      - od_nominal_mm, od_tolerance_mm, od_formatted  
      - thickness_mm, thickness_tolerance_mm, thickness_formatted
      - burst_pressure_mpa, burst_pressure_formatted
    """
    res = dict(result)  # shallow copy
    dimensions = res.get('dimensions', {})

    # ID nominal and tolerance
    id_nom = res.get('id_nominal_mm')
    id_tol = res.get('id_tolerance_mm')
    # if id_nom missing try to parse from other keys
    if id_nom is None:
        id_candidates = [dimensions.get(k) for k in ('id1','id','ID','nominal_id_mm') if dimensions.get(k) is not None]
        if id_candidates:
            try:
                id_nom = float(id_candidates[0])
                res['id_nominal_mm'] = id_nom
            except Exception:
                id_nom = None
    # Ensure numeric types
    try:
        if id_nom is not None:
            id_nom = float(id_nom)
    except Exception:
        id_nom = None
    try:
        if id_tol is not None:
            id_tol = float(id_tol)
    except Exception:
        id_tol = None
    res['id_nominal_mm'] = id_nom
    res['id_tolerance_mm'] = id_tol
    res['id_formatted'] = format_tolerance(id_nom, id_tol) or "N/A"

    # OD nominal and tolerance
    od_nom = dimensions.get('od1') or res.get('od_nominal_mm') or res.get('od')
    od_tol = res.get('od_tolerance_mm')
    try:
        if od_nom is not None and od_nom != 'Not Found':
            od_nom = float(od_nom)
    except Exception:
        od_nom = None
    try:
        if od_tol is not None:
            od_tol = float(od_tol)
    except Exception:
        od_tol = None
    res['od_nominal_mm'] = od_nom
    res['od_tolerance_mm'] = od_tol
    res['od_formatted'] = format_tolerance(od_nom, od_tol) or "N/A"

    # Thickness: prefer explicit, else compute from OD/ID (guarded against table overwrites)
    thickness = res.get('thickness_mm') or res.get('thickness')
    thickness_tol = res.get('thickness_tolerance_mm')
    provenance_thickness = res.get('thickness_source')
    try:
        if thickness is not None and thickness != 'Not Found':
            thickness = float(thickness)
    except Exception:
        thickness = None
    
    # Only compute thickness if not already set by authoritative table
    if thickness is None and od_nom is not None and id_nom is not None and provenance_thickness != 'TABLE-4/8 Grade-1 authoritative':
        try:
            thickness = round((od_nom - id_nom) / 2.0, 3)
            # set a reasonable default tolerance if not supplied by table
            if thickness_tol is None:
                thickness_tol = 0.25
            logging.debug(f"ensure_result_fields: Computed thickness from OD/ID for {res.get('part_number','?')}: {thickness} mm (provenance={provenance_thickness})")
        except Exception:
            thickness = None
    elif provenance_thickness == 'TABLE-4/8 Grade-1 authoritative':
        logging.debug(f"ensure_result_fields: Thickness for {res.get('part_number','?')} is authoritative from table: {thickness} mm ± {thickness_tol} mm")
    
    res['thickness_mm'] = thickness
    res['thickness_tolerance_mm'] = thickness_tol
    res['thickness_formatted'] = format_tolerance(thickness, thickness_tol) or "N/A"

    # Burst pressure
    burst = res.get('burst_pressure_mpa')
    try:
        if burst is not None:
            burst = float(burst)
    except Exception:
        burst = None
    res['burst_pressure_mpa'] = burst
    res['burst_pressure_formatted'] = f"{burst:.2f} MPa" if burst is not None else "N/A"

    return res

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def generate_corrected_excel_sheet(analysis_results, dimensions, coordinates):
    """
    Generate enhanced Excel output with proper formatting and data validation.
    
    Args:
        analysis_results (dict): Results from drawing analysis
        dimensions (dict): Extracted dimensions
        coordinates (list): Extracted coordinate points
    
    Returns:
        BytesIO: Excel file data
    """
    try:
        # Define column structure with proper formatting
        columns = [
            'child part',                                         # Row 1: Original format
            'child quantity',                                     # Row 1: Original format
            'CHILD PART',                                        # Row 2: Uppercase format
            'CHILD PART DESCRIPTION',                            # Description
            'CHILD PART QTY',                                    # Quantity
            'SPECIFICATION',                                     # Combined standard+grade
            'MATERIAL',                                         # From database lookup
            'POLYMER TYPE',                                      # For ASTM D2000 materials
            'REINFORCEMENT',                                     # Additional info
            'RINGS',                                            # Rings information
            'ID1 AS PER 2D (MM)',                              # First ID measurement
            'ID TOLERANCE (MM)',                                # MPAPS F-6032 ID tolerance
            'ID2 AS PER 2D (MM)',                              # Second ID measurement
            'OD1 AS PER 2D (MM)',                              # First OD measurement
            'OD TOLERANCE (MM)',                                # MPAPS F-6032 OD tolerance
            'OD2 AS PER 2D (MM)',                              # Second OD measurement
            'THICKNESS AS PER 2D (MM)',                        # Direct thickness
            'WALL THICKNESS TOLERANCE (MM)',                   # Wall thickness tolerance from standards
            'THICKNESS AS PER ID OD DIFFERENCE',               # Calculated thickness
            'BURST PRESSURE (MPA)',                           # MPAPS F-6032 burst pressure
            'CENTERLINE LENGTH AS PER 2D (MM)',                # From drawing
            'DEVELOPMENT LENGTH AS PER CO-ORDINATE (MM)',      # Calculated length
            'BURST PRESSURE AS PER 2D (BAR)',                  # From drawing
            'BURST PRESSURE AS PER WORKING PRESSURE (4XWP) (BAR)', # Calculated
            'VOLUME AS PER 2D MM3',                            # Volume in mm³
            'WEIGHT AS PER 2D KG',                             # Weight from drawing
            'COLOUR AS PER DRAWING',                           # Color specification
            'ADDITIONAL REQUIREMENT',                          # Notes and requirements
            'OUTSOURCE',                                       # Outsourcing details
            'REMARK'                                          # Generated remarks
        ]

        # Get part number and clean it
        part_number = str(analysis_results.get('part_number', '')).strip()
        if not part_number or part_number == 'Not Found':
            part_number = "Unknown Part"

        # Get description and clean it
        description = str(analysis_results.get('description', '')).strip()
        if not description or description == 'Not Found':
            description = "No Description Available"

        # Format specification string
        standard = analysis_results.get('standard', 'Not Found')
        grade = analysis_results.get('grade', 'Not Found')
        specification = f"{standard}"
        if grade != 'Not Found':
            specification += f" {grade}"

        # Post-process results to ensure fields are properly formatted
        # **PATCH 2B**: Ensure MPAPS values have been applied (populate thickness, tolerances, etc.)
        try:
            from mpaps_utils import process_mpaps_dimensions
            analysis_results = process_mpaps_dimensions(analysis_results or {})
            logging.debug("process_mpaps_dimensions called successfully before ensure_result_fields")
        except Exception as e:
            logging.debug(f"process_mpaps_dimensions failed inside excel generator: {e}", exc_info=True)

        analysis_results = ensure_result_fields(analysis_results)
        
        # Debug: Log value provenance before Excel creation
        logging.info("DEBUG BEFORE EXCEL: part_number=%s, id_nominal_mm=%s, id_tolerance_mm=%s, od_nominal_mm=%s, thickness_mm=%s, thickness_tolerance_mm=%s, dimension_source=%s, thickness_source=%s",
                     analysis_results.get('part_number'),
                     analysis_results.get('id_nominal_mm'),
                     analysis_results.get('id_tolerance_mm'),
                     analysis_results.get('od_nominal_mm'),
                     analysis_results.get('thickness_mm'),
                     analysis_results.get('thickness_tolerance_mm'),
                     analysis_results.get('dimension_source'),
                     analysis_results.get('thickness_source'))
        
        # Get formatted values with proper handling of N/A
        thickness_calculated = analysis_results.get('thickness_formatted', "N/A")
        
        # Calculate development length with robust error handling
        if coordinates:
            try:
                development_length_mm, is_fallback, reason = safe_development_length(coordinates)
                
                # Add sanity check: reject unrealistic development lengths (>20 meters = 20,000 mm)
                if development_length_mm is not None:
                    if development_length_mm <= 0 or development_length_mm > 20000:
                        logging.warning(f"Unrealistic development length {development_length_mm}mm; rejecting and falling back")
                        development_length_mm = None
                
                development_length = f"{development_length_mm:.2f} mm" if development_length_mm is not None else "Not Found"
                if is_fallback:
                    logging.warning("Development length calculation used fallback: %s", reason)
            except ValueError as e:
                logging.warning(f"safe_development_length failed: {e} - falling back to centerline")
                development_length = "Not Found"
        else:
            development_length = "Not Found"
            logger.warning("No coordinates available for development length calculation")

        # Calculate burst pressure from working pressure
        burst_pressure_calc = "Not Found"
        if analysis_results.get('working_pressure') not in ['Not Found', None]:
            try:
                wp = float(str(analysis_results['working_pressure']).replace(',', '.'))
                burst_pressure_calc = f"{(wp * 4):.1f}"
            except (ValueError, TypeError):
                pass

        # Get wall thickness from either direct measurement or calculated
        wall_thickness = analysis_results.get('wall_thickness')  # From Grade 1B/1BF rules
        if wall_thickness is None:
            wall_thickness = dimensions.get('thickness', 'Not Found')
        if wall_thickness == 'Not Found' and thickness_calculated != 'Not Found':
            wall_thickness = thickness_calculated

        # At this point analysis_results has been passed through ensure_result_fields()
        # so use that canonical dict (snake_case keys) to populate Excel columns.
        res = analysis_results  # canonical normalized dict from ensure_result_fields

        # Debug: log the normalized fields we rely on
        logging.info("Normalized fields: id_formatted=%s, od_formatted=%s, thickness_formatted=%s, burst_pressure_formatted=%s",
                     res.get('id_formatted'), res.get('od_formatted'), res.get('thickness_formatted'), res.get('burst_pressure_formatted'))

        # Prepare display strings (use 'N/A' when missing)
        id_display = res.get('id_formatted') or "N/A"
        od_display = res.get('od_formatted') or "N/A"
        thickness_display = res.get('thickness_formatted') or "N/A"
        burst_display = res.get('burst_pressure_formatted') or "N/A"

        # Build the row data dictionary using the display strings and other normalized fields
        row_data = {
            'child part': part_number.lower(),
            'child quantity': "1",
            'CHILD PART': part_number.upper(),
            'CHILD PART DESCRIPTION': description,
            'CHILD PART QTY': "1",
            'SPECIFICATION': specification,
            'MATERIAL': res.get('material', 'Not Found'),
            'POLYMER TYPE': res.get('polymer_type', 'Not Applicable'),
            'REINFORCEMENT': res.get('reinforcement', 'Not Found'),
            'RINGS': res.get('rings', 'Not Found'),
            # Use the formatted display strings here
            'ID1 AS PER 2D (MM)': id_display,
            'ID TOLERANCE (MM)': format_tolerance(None, res.get('id_tolerance_mm')) or 'N/A',
            'ID2 AS PER 2D (MM)': id_display,
            'OD1 AS PER 2D (MM)': od_display,
            'OD TOLERANCE (MM)': format_tolerance(None, res.get('od_tolerance_mm')) or 'N/A',
            'OD2 AS PER 2D (MM)': od_display,
            'THICKNESS AS PER 2D (MM)': thickness_display,
            'WALL THICKNESS TOLERANCE (MM)': format_tolerance(None, res.get('thickness_tolerance_mm')) or 'N/A',
            'THICKNESS AS PER ID OD DIFFERENCE': thickness_display,
            'BURST PRESSURE (MPA)': burst_display,
            'CENTERLINE LENGTH AS PER 2D (MM)': dimensions.get('centerline_length', 'Not Found'),
            'DEVELOPMENT LENGTH AS PER CO-ORDINATE (MM)': development_length,
            'BURST PRESSURE AS PER 2D (BAR)': res.get('burst_pressure', 'Not Found'),
            'BURST PRESSURE AS PER WORKING PRESSURE (4XWP) (BAR)': burst_pressure_calc,
            'VOLUME AS PER 2D MM3': res.get('volume_mm3', 'Not Found'),
            'WEIGHT AS PER 2D KG': res.get('weight', 'Not Found'),
            'COLOUR AS PER DRAWING': res.get('color', 'Not Found'),
            'ADDITIONAL REQUIREMENT': "CUTTING & CHECKING FIXTURE COST TO BE ADDED. Marking cost to be added.",
            'OUTSOURCE': "",
            'REMARK': ""
        }

        # Generate remarks based on validation (keep your existing logic)
        remarks = []
        if standard.startswith('MPAPS F 1'):
            remarks.append('Drawing specifies MPAPS F 1, considered as MPAPS F 30.')
        grade = res.get('grade', '')
        from mpaps_utils import is_grade_1bf
        if is_grade_1bf(grade):
            if res.get('wall_thickness'):
                remarks.append(f"Using Grade 1B/1BF wall thickness: {res['wall_thickness']} mm")
                if res.get('wall_thickness_tolerance'):
                    remarks.append(f"Wall thickness tolerance: {res['wall_thickness_tolerance']}")
            if res.get('od_reference'):
                remarks.append(f"Using Grade 1B/1BF reference OD: {res['od_reference']} mm")

        # ID/OD mismatch checks (use dimensions dict)
        id1 = dimensions.get('id1', 'Not Found')
        id2 = dimensions.get('id2', 'Not Found')
        if id1 != 'Not Found' and id2 != 'Not Found' and id1 != id2:
            remarks.append('THERE IS MISMATCH IN ID 1 & ID 2')
        od1 = dimensions.get('od1', 'Not Found')
        od2 = dimensions.get('od2', 'Not Found')
        if od1 != 'Not Found' and od2 != 'Not Found' and od1 != od2:
            remarks.append('THERE IS MISMATCH IN OD 1 & OD 2')

        row_data['REMARK'] = ' '.join(remarks) if remarks else 'No specific remarks.'

        # Create DataFrame with formatted data
        df = pd.DataFrame([row_data], columns=columns)

        # Debug log to see what's going into Excel (log actual columns present)
        cols_to_log = [c for c in ['ID1 AS PER 2D (MM)', 'OD1 AS PER 2D (MM)', 'THICKNESS AS PER 2D (MM)', 'BURST PRESSURE (MPA)'] if c in df.columns]
        logging.info("Prepared DataFrame head for Excel write:\n%s", df[cols_to_log].to_string(index=False))

        # Create Excel writer with enhanced formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='FETCH FROM DRAWING', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['FETCH FROM DRAWING']
            
            # Define styles
            header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            header_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting
            for col_idx, column in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Auto-fit column width
                max_length = max(
                    len(str(column)),
                    len(str(df.iloc[0][column]))
                )
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # Apply borders to data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating Excel sheet: {e}", exc_info=True)
        # Create a minimal error sheet
        error_df = pd.DataFrame([{
            'child part': 'ERROR',
            'REMARK': f'Excel generation failed: {str(e)}'
        }], columns=columns)
        error_output = io.BytesIO()
        error_df.to_excel(error_output, sheet_name='FETCH FROM DRAWING', index=False)
        error_output.seek(0)
        return error_output